import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def comparar_planilhas(planilha_base, planilha_atualizada, resultado):
    # Carregar as planilhas
    
    # Carrega planilha ANTIGA
    df1 = pd.read_excel(planilha_base, engine='openpyxl')

    # Carrega planilha NOVA
    df2 = pd.read_excel(planilha_atualizada, engine='openpyxl')

    # Mescla os dois dataframes
    merged = pd.merge(df1, df2, on=list(df1.columns), how='outer', indicator=True)
    
    # Salva a nova planilha
    merged.to_excel(resultado, index=False)
    
    # Adiciona formatação condicional
    wb = load_workbook(resultado)
    ws = wb.active
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[-1].value == 'left_only':
            for cell in row:
                cell.fill = red_fill
        elif row[-1].value == 'right_only':
            for cell in row:
                cell.fill = green_fill
    
    wb.save(resultado)

if __name__ == "__main__":
    planilha_base = "sheet1.xlsx"
    planilha_atualizada = "sheet2.xlsx"
    resultado = "differences.xlsx"

    comparar_planilhas(planilha_base, planilha_atualizada, resultado)
